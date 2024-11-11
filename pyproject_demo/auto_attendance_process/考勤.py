import openpyxl # openpyxl打开xlsx文件，行列下标直接从1开始，不是0
import xlrd # xlrd 库在较新版本中不再支持读取 .xlsx 格式的 Excel 文件。xlrd 现在只支持 .xls 格式
"""
迟到早退
    0<time<10 扣20
    10<time 扣50
只打一次
    扣 20
不打卡
    扣1.5倍工资
调休
    当月天数够，不扣；不够，扣1.5倍工资
"""
def rans(num): # 处理数据方法
    temp_rs = [0,0,0,0] # 迟到10分钟 迟到10分钟以上 只打一次卡 旷工
    for i in range(2,31):
        t = sheet.cell(row=num, column=i).value
        # print(sheet.cell_value(num, 0))
        all_time = str(t).split('\n') if t != 0 else []
        if all_time == ['None']:
            all_time = []
        count = len(all_time)
        # print(all_time)
        # print(count)
        if count == 0: # 没有打卡
            temp_rs[3] += 1 # 记录一次旷工
        else: # 已经打卡
            start_time = all_time[0]
            t_start = transform(start_time)
            if count == 1: # 只打了一次卡
                temp_rs[2] += 1 # 记录只打一次卡
                c = compare_time(t_start)
                statistics(c, temp_rs)
            else: # 打了多次
                end_time = all_time[-1]
                t_end = transform(end_time)
                if t_end[0] - t_start[0] < 4:
                    temp_rs[2] += 1  # 记录只打一次卡
                    if t_end[0] < 12: # 判断多次打卡是不是都是上午卡
                        c = compare_time(t_start)
                    else:
                        c = compare_time(t_end)
                    statistics(c, temp_rs)
                else:
                    a = compare_time(t_start)
                    p = compare_time(t_end)
                    statistics(a,temp_rs)
                    statistics(p,temp_rs)
    return temp_rs

# def transform(tmp_time): # 处理时间
#     hour, minute = tmp_time.split(':')
#     sum_minute = int(hour)*60 + int(minute)
#     return [int(hour),int(minute),sum_minute]
def transform(tmp_time):  # 处理时间
    # 检查 tmp_time 是否为字符串且包含 ':'
    if tmp_time and isinstance(tmp_time, str) and ':' in tmp_time:
        hour, minute = tmp_time.split(':')
        sum_minute = int(hour) * 60 + int(minute)
        return [int(hour), int(minute), sum_minute]
    else:
        # 返回默认值（例如 [0, 0, 0]），表示无效输入
        return [0, 0, 0]

def compare_time(tmp_time): # 判断是否有迟到或早退
    if tmp_time[0] < 12: # 打的上午卡
        c = tmp_time[2] - 9*60
    else: # 签退
        c = 18*60 - tmp_time[2]
    return c if c > 0 else 0

def statistics(c,temp_rs): # 判断迟到或早退了几分钟
    if c > 0 and c <= 10:
        temp_rs[0] += 1
    elif c > 10:
        temp_rs[1] += 1



if __name__ == '__main__':
    excel = openpyxl.load_workbook('/Users/chenzhiyong/PycharmProjects/lilnero/pyproject_demo/auto_attendance_process/kq.xlsx') # 必须是绝对路径
    sheet = excel.worksheets[0]
    # for i in range(1,sheet.nrows): nrows是xlrd库里的方法，不能用
    #     one_rs = rans(i)
    for i in range(2, sheet.max_row + 1):  # sheet.max_row 获取最大行数
        one_rs = rans(i)
        print(f'姓名：{sheet.cell(row=i, column=1).value}，迟到十分钟以内：{one_rs[0]}次，迟到十分钟以上：{one_rs[1]}次，单次打卡：{one_rs[2]}次，旷工：{one_rs[3]}次')
