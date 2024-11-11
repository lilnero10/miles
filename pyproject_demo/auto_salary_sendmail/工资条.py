from email.header import Header
from openpyxl import load_workbook
import smtplib
from email.mime.text import MIMEText

wb = load_workbook('/pyproject_demo/auto_salary_sendmail/salary.xlsx')

sheet1 = wb.active

count = 0
table_header = ''
for row in sheet1.iter_rows():
    if count == 0:
        table_header = f'''
                <tr>
                    <td>{row[1].value}</td>
                    <td>{row[2].value}</td>
                    <td>{row[3].value}</td>
                    <td>{row[4].value}</td>
                    <td>{row[9].value}</td>
                </tr>
                '''
        count += 1 # 避免读到表头
        continue
    msg = f'''
                <tr>
                    <td>{row[1].value}</td>
                    <td>{row[2].value}</td>
                    <td>{row[3].value}</td>
                    <td>{row[4].value}</td>
                    <td>{row[9].value}</td>
                </tr>
                '''
    count += 1
    to_addr = row[9].value

    # 发送邮件
    smtp_obj = smtplib.SMTP('smtp.qq.com')
    # 登录邮箱
    smtp_obj.login('2357577930@qq.com','') # 密码在QQ邮箱设置获取
    # 定义发送的内容
    message_body = MIMEText('<table border="1">'+table_header+msg+'</table>', 'html', 'utf-8')
    # 定义从哪发来的邮件
    message_body['From'] = Header('测试部门','utf-8')
    # 邮件主题
    message_body['Subject'] = Header('第一次发送邮件','utf-8')
    # 发送
    smtp_obj.sendmail('2357577930@qq.com',[to_addr],message_body.as_string())
    print(f'已成功发送邮件给名为{row[1].value}邮箱地址{to_addr}')