import pandas as pd
import openpyxl

# 读取Excel文件
df = pd.read_excel('')

# 打开指定的 Excel 文件，然后在指定的工作表中取消所有合并的单元格，并保存修改后的 Excel 文件
def unmerge_cells(df, sheet1):
    sj = openpyxl.load_workbook(df)
    sheet = sj[sheet1]

    for merged_cell in sheet.merged_cells.ranges:
        sheet.unmerge_cells(str(merged_cell))

    sj.save(df)

# 删除小票号为空的行
df = df.dropna(subset=['小票号'])
# 将结果输出到csv文件
df.to_csv('处理后数据2.csv', index=False)
